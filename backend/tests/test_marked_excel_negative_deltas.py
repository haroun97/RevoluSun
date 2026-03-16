"""Verify that the marked Excel highlights exactly the rows with negative deltas.

This test is deliberately end-to-end on the Excel files only:
- It recomputes negative deltas per sheet from the *original* workbook.
- It then checks that the corresponding cells in the *marked* workbook
  (`Messdaten_Nürnberg_2024-2026_marked.xlsx`) are filled in red.
"""
from pathlib import Path

import openpyxl
import pandas as pd


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DOC_DIR = PROJECT_ROOT / "document"
ORIGINAL_XLSX = DOC_DIR / "Messdaten_Nürnberg_2024-2026.xlsx"
MARKED_XLSX = DOC_DIR / "Messdaten_Nürnberg_2024-2026_marked.xlsx"

TIMESTAMP_COLS = ("timestamp", "Timestamp", "Zeit", "Datum", "Date", "time", "datetime")
VALUE_COLS = ("value", "Value", "Wert", "Reading", "raw_value", "kwh", "cumulative")


def _detect_timestamp_column(df: pd.DataFrame) -> str | None:
    for c in TIMESTAMP_COLS:
        if c in df.columns:
            return c
    for c in df.columns:
        name = str(c).lower()
        if "date" in name or "zeit" in name or "time" in name:
            return c
    return None


def _detect_value_column(df: pd.DataFrame) -> str | None:
    for c in VALUE_COLS:
        if c in df.columns:
            return c
    for c in df.columns:
        name = str(c).lower()
        if "value" in name or "wert" in name or "kwh" in name or "reading" in name:
            return c
    for c in df.columns:
        if pd.api.types.is_numeric_dtype(df[c]):
            return c
    return None


def _compute_negative_rows(df: pd.DataFrame, conv: float) -> pd.DataFrame:
    ts_col = _detect_timestamp_column(df)
    val_col = _detect_value_column(df)
    if not ts_col or not val_col:
        return pd.DataFrame()

    work = df[[ts_col, val_col]].copy()
    work.columns = ["timestamp_raw", "value_raw"]
    work["timestamp"] = pd.to_datetime(work["timestamp_raw"], errors="coerce")
    work["value"] = pd.to_numeric(work["value_raw"], errors="coerce")
    work = work.dropna(subset=["timestamp", "value"])
    if work.empty:
        return pd.DataFrame()

    work["cumulative"] = work["value"] * conv
    work["original_row"] = work.index
    work = work.sort_values("timestamp").reset_index(drop=True)
    work["prev_cumulative"] = work["cumulative"].shift(1)
    work["delta"] = work["cumulative"] - work["prev_cumulative"]

    neg = work[work["delta"] < 0].copy()
    if neg.empty:
        return pd.DataFrame()
    neg["value_column"] = val_col
    return neg[["original_row", "timestamp", "cumulative", "delta", "value_column"]]


def _header_map(ws) -> dict[str, str]:
    header = next(ws.iter_rows(min_row=1, max_row=1))
    mapping: dict[str, str] = {}
    for cell in header:
        if cell.value is None:
            continue
        mapping[str(cell.value)] = cell.column_letter
    return mapping


def _is_red_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    rgb = (fill.start_color.rgb or "").upper()
    return rgb in {"FFFFC0CB", "FFC0CBFF"} or "FFC0CB" in rgb


def test_marked_excel_highlights_negative_deltas():
    assert ORIGINAL_XLSX.is_file(), f"Original Excel not found at {ORIGINAL_XLSX}"
    assert MARKED_XLSX.is_file(), f"Marked Excel not found at {MARKED_XLSX}"

    # Load workbooks
    xl = pd.ExcelFile(ORIGINAL_XLSX, engine="openpyxl")
    wb_marked = openpyxl.load_workbook(MARKED_XLSX)

    BUILDING_SHEETS = {"Summenzähler", "Summe", "Building", "building_total"}

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        conv = 50.0 if sheet_name in BUILDING_SHEETS else 1.0
        neg = _compute_negative_rows(df, conv)

        ws = wb_marked[sheet_name]
        header = _header_map(ws)

        if neg.empty:
            # No negatives: there should be no red cells in any value-like column
            red_cells = [
                cell
                for row in ws.iter_rows(min_row=2)
                for cell in row
                if _is_red_fill(cell)
            ]
            assert not red_cells, f"Unexpected red cells in sheet {sheet_name}"
            continue

        # For each negative row, check that the corresponding value cell is red.
        for _, row in neg.iterrows():
            excel_row = int(row["original_row"]) + 2  # header row is 1
            col_name = row["value_column"]
            col_letter = header.get(col_name)
            assert col_letter, f"Column {col_name} not found in sheet {sheet_name}"
            cell = ws[f"{col_letter}{excel_row}"]
            assert _is_red_fill(cell), (
                f"Expected red fill for negative delta in sheet {sheet_name}, "
                f"row {excel_row}, column {col_name}"
            )

