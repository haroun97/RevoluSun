#!/usr/bin/env python3
"""
Export a copy of the source Excel file with cells that contribute to
negative deltas highlighted in red.

This is a QA/diagnostic tool only; it does not touch the database.

Usage (from backend/):

  .venv/bin/python scripts/mark_negative_deltas_in_excel.py

It will:
- read `../document/Messdaten_Nürnberg_2024-2026.xlsx`
- detect timestamp / value columns per sheet
- recompute cumulative values and deltas per sheet
- for every row where delta < 0, color the *value* cell red
- write a new workbook next to the original, e.g.
  `../document/Messdaten_Nürnberg_2024-2026_marked.xlsx`
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill


EXCEL_REL_PATH = Path("..") / "document" / "Messdaten_Nürnberg_2024-2026.xlsx"

TIMESTAMP_COLS = ("timestamp", "Timestamp", "Zeit", "Datum", "Date", "time", "datetime")
VALUE_COLS = ("value", "Value", "Wert", "Reading", "raw_value", "kwh", "cumulative")


def detect_timestamp_column(df: pd.DataFrame) -> str | None:
    for c in TIMESTAMP_COLS:
        if c in df.columns:
            return c
    for c in df.columns:
        name = str(c).lower()
        if "date" in name or "zeit" in name or "time" in name:
            return c
    return None


def detect_value_column(df: pd.DataFrame) -> str | None:
    for c in VALUE_COLS:
        if c in df.columns:
            return c
    for c in df.columns:
        name = str(c).lower()
        if "value" in name or "wert" in name or "kwh" in name or "reading" in name:
            return c
    for c in df.columns:
        if pd.api.types.is_numeric_dtype(df[c]):
            return c
    return None


def compute_negative_rows(df: pd.DataFrame, conversion_factor: float) -> pd.DataFrame:
    """
    Given a raw DataFrame from read_excel, return rows where the cumulative
    value decreases after sorting by timestamp.

    The returned DataFrame has columns:
      - original_row (0-based index from pandas, corresponding to Excel row)
      - timestamp
      - prev_timestamp
      - cumulative
      - prev_cumulative
      - delta
      - value_column (name of the column to highlight)
    """
    ts_col = detect_timestamp_column(df)
    val_col = detect_value_column(df)
    if not ts_col or not val_col:
        return pd.DataFrame()

    work = df[[ts_col, val_col]].copy()
    work.columns = ["timestamp_raw", "value_raw"]
    work["timestamp"] = pd.to_datetime(work["timestamp_raw"], errors="coerce")
    work["value"] = pd.to_numeric(work["value_raw"], errors="coerce")
    work = work.dropna(subset=["timestamp", "value"])
    if work.empty:
        return pd.DataFrame()

    work["cumulative"] = work["value"] * conversion_factor
    work["original_row"] = work.index

    work = work.sort_values("timestamp").reset_index(drop=True)
    work["prev_cumulative"] = work["cumulative"].shift(1)
    work["prev_timestamp"] = work["timestamp"].shift(1)
    work["delta"] = work["cumulative"] - work["prev_cumulative"]

    neg = work[work["delta"] < 0].copy()
    if neg.empty:
        return pd.DataFrame()

    neg["value_column"] = val_col
    return neg[
        [
            "original_row",
            "timestamp",
            "prev_timestamp",
            "cumulative",
            "prev_cumulative",
            "delta",
            "value_column",
        ]
    ]


def build_column_letter_map(ws) -> dict[str, str]:
    """
    Map header text -> column letter for a worksheet, based on the first row.
    Header cells that are empty are ignored.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    mapping: dict[str, str] = {}
    for cell in header_row:
        if cell.value is None:
            continue
        mapping[str(cell.value)] = cell.column_letter
    return mapping


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a marked copy of the Excel with negative deltas highlighted.")
    parser.add_argument(
        "--excel-path",
        type=str,
        default=str(EXCEL_REL_PATH),
        help="Path to the source Excel file (default: ../document/Messdaten_Nürnberg_2024-2026.xlsx).",
    )
    parser.add_argument(
        "--output-path",
        type=str,
        default="",
        help="Optional explicit output path for the marked Excel. "
        "If not set, '_marked' is appended before the extension.",
    )
    args = parser.parse_args()

    src_path = Path(args.excel_path).resolve()
    if not src_path.is_file():
        print(f"Source Excel not found: {src_path}")
        return

    if args.output_path:
        out_path = Path(args.output_path).resolve()
    else:
        out_path = src_path.with_name(src_path.stem + "_marked" + src_path.suffix)

    print(f"Reading source Excel: {src_path}")

    # Use pandas to detect negative rows per sheet
    xl = pd.ExcelFile(src_path, engine="openpyxl")

    BUILDING_SHEETS = {"Summenzähler", "Summe", "Building", "building_total"}

    negatives_by_sheet: dict[str, pd.DataFrame] = {}
    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        conv = 50.0 if sheet_name in BUILDING_SHEETS else 1.0
        neg = compute_negative_rows(df, conv)
        if not neg.empty:
            negatives_by_sheet[sheet_name] = neg
            print(f"{sheet_name}: {len(neg)} negative rows will be highlighted")
        else:
            print(f"{sheet_name}: no negative rows detected")

    if not negatives_by_sheet:
        print("No negative deltas found in any sheet; nothing to mark.")
        return

    # Load workbook with openpyxl to apply cell styles
    wb = openpyxl.load_workbook(src_path)
    red_fill = PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid")

    for sheet_name, neg_df in negatives_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_map = build_column_letter_map(ws)
        if neg_df.empty:
            continue

        # Assume pandas treated the first Excel row as header, so data starts at row 2.
        # original_row is 0-based for the first data row -> Excel row = original_row + 2
        for _, row in neg_df.iterrows():
            excel_row = int(row["original_row"]) + 2
            value_col_name = row["value_column"]
            col_letter = header_map.get(value_col_name)
            if not col_letter:
                # Fallback: do nothing if we cannot find the header
                continue
            cell_ref = f"{col_letter}{excel_row}"
            cell = ws[cell_ref]
            cell.fill = red_fill

    wb.save(out_path)
    print(f"Wrote marked workbook to: {out_path}")


if __name__ == "__main__":
    main()

